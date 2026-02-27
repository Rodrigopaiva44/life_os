"""
init_core.py
============
Entrypoint principal do Life_OS.
Gera o arquivo Life_OS_V1.xlsx com todas as abas da arquitetura V1.

Abas geridas aqui (via formatar_aba genérico):
  Facul_Horario, Facul_Disciplinas,
  Fin_FluxoCaixa, Fin_Investimentos, Saude_Dieta

Aba gerida pelo módulo dedicado (sheets.work_projetos):
  Work_Projetos  ← Data Validation + Conditional Formatting + AutoFilter
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from sheets.work_projetos import build_work_projetos_sheet


# ── Formatador genérico para abas simples ─────────────────────────────────────

def formatar_aba(ws, df: pd.DataFrame) -> None:
    """Aplica formatação Enterprise aos cabeçalhos e auto-ajusta largura de coluna."""
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", color="FFFFFF", bold=True)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[col[0].column_letter].width = max_length + 5

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"


# ── Definição das abas ────────────────────────────────────────────────────────

def _build_sheets_data() -> dict[str, pd.DataFrame]:
    df_horario = pd.DataFrame({
        "Horario":  ["M1 (07:00-07:50)", "M2 (07:50-08:40)", "M3 (08:50-09:40)",
                     "M4 (09:40-10:30)", "T1 (12:30-13:20)"],
        "Segunda":  ["", "", "LEITURA", "ACADEMIA", "TRABALHO"],
        "Terca":    ["", "", "LEITURA", "ACADEMIA", "TRABALHO"],
        "Quarta":   ["", "", "LEITURA", "ACADEMIA", "TRABALHO"],
        "Quinta":   ["", "", "LEITURA", "ACADEMIA", "TRABALHO"],
        "Sexta":    ["", "", "LEITURA", "ACADEMIA", "TRABALHO"],
    })

    df_materias = pd.DataFrame(columns=[
        "Materia", "Professor", "Email_Professor", "Data_P1", "Data_P2",
        "Data_Final", "Faltas", "Status_Materia", "Observacoes",
    ])

    df_fluxo_caixa = pd.DataFrame(columns=[
        "Data_Transacao", "Tipo", "Categoria", "Descricao",
        "Valor_BRL", "Conta_Origem", "Status",
    ])

    df_investimentos = pd.DataFrame(columns=[
        "Ticker", "Mercado", "Quantidade", "Preco_Medio",
        "Cotacao_Atual", "Valor_Alocado", "Lucro_Perda_Perc",
    ])

    df_saude = pd.DataFrame(columns=[
        "Refeicao", "Alimento", "Quantidade", "Carboidratos_g",
        "Proteinas_g", "Gorduras_g", "Calorias_kcal",
    ])

    return {
        "Facul_Horario":      df_horario,
        "Facul_Disciplinas":  df_materias,
        "Fin_FluxoCaixa":     df_fluxo_caixa,
        "Fin_Investimentos":  df_investimentos,
        "Saude_Dieta":        df_saude,
    }


# ── Entrypoint ────────────────────────────────────────────────────────────────

def build_life_os(arquivo: str = "Life_OS_V1.xlsx") -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # Abas genéricas
    for nome_aba, df in _build_sheets_data().items():
        ws = wb.create_sheet(title=nome_aba)
        formatar_aba(ws, df)

    # Work_Projetos: módulo dedicado com DV + Conditional Formatting + AutoFilter
    build_work_projetos_sheet(wb)

    wb.save(arquivo)
    print(f"[OK] {arquivo} gerado com sucesso. Arquitetura V1 inicializada.")


if __name__ == "__main__":
    build_life_os()
